import openpyxl as oxl
import pprint
wb = oxl.load_workbook('./Book1.xlsx')
l_sheet = wb.sheetnames
def xacdinh(sheet_name):
    a = 1
    b = 1
    sheet = wb[sheet_name]
    run = True
    while(run == True):
        cell = sheet.cell(row=a, column=b)
        if cell.value != None and b == 1:
            a += 1
            continue
        elif cell.value == None and b == 1:
            b += 1
            a -= 1
            continue
        if cell.value != None and b != 1:
            b += 1
            continue
        elif cell.value == None and b != 1:
            run = False
    return a , b - 1
def xacdinhSDT(sheet_name, b):
    sheet = wb[sheet_name]
    for i in range(1,b + 1):
        cell = sheet.cell(row=1, column=i)
        if cell.value in ['SDT','DTHOAI','SODT']:
            return i
        elif cell.value == None:
            return 0

for n_sheet in l_sheet:
    sheet_name = n_sheet
    sheet = wb[sheet_name]
    a , b= xacdinh(sheet_name)
    SDT_pos = xacdinhSDT(sheet_name,b)
    print(f'-- Nhap lieu cho bang {sheet_name} --')
    for i in range(1+1,a+1):
        list = []
        for j in range(1,b+1):
            cell = sheet.cell(row=i, column=j)
            if j == SDT_pos:
                cell.value = str(cell.value)
                cell.value = '0' + cell.value
            if type(cell.value) not in [int,float,None]:
                cell.value = str(cell.value)
            list.append(cell.value)
        str1 = f"INSERT INTO {sheet_name} VALUES ({list[0:b]}) "
        str1 = str1.replace("[","").replace("]","").replace(" 00:00:00","").replace("-","/")
        str1 = str1.replace("'Null'","Null").replace("'NULL'","NULL")
        print(str1)